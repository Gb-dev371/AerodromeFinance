from openpyxl import load_workbook
from collections import defaultdict
import os
from pathlib import Path

def get(pool_name, gauge_address):
    # wb = load_workbook(f'GaugesDataCurve\SYMBOL={pool_name}ADDRESS={gauge_address}.xlsx', data_only=True)
    # return wb.active

    wb = load_workbook(f'curve\GaugesDataCurve\SYMBOL={pool_name}ADDRESS={gauge_address}.xlsx', data_only=True)

    # Planilha com os dados de Amount e Gauge Address
    ws = wb.active
    return ws


def bribe_analyzer(gauge_file_path, round_votium):
    # Carregar o workbook original e o novo com os dados do pool
    wb = load_workbook(gauge_file_path, data_only=True)
    wb_pools = load_workbook('curve/curve_gauge_data_2.xlsx', data_only=True)

    # Planilha com os dados de Amount e Gauge Address
    ws = wb['Sheet']

    # Planilha com os dados de Pool Name e Gauge Address
    ws_pools = wb_pools.active

    # Dicionário para mapear Gauge Address a Pool Name
    gauge_to_poolname = {}

    # Lendo dados de Pool Name e Gauge Address
    for row in ws_pools.iter_rows(min_row=2, values_only=True):
        gauge_address = (row[1]).lower()  # Coluna B no arquivo novo
        pool_name = row[0]      # Coluna A no arquivo novo
        gauge_to_poolname[gauge_address] = pool_name # fazendo algo que não precisaria se fosse json

    # Dicionário para armazenar a soma dos valores de Amount $ por gauge
    gauge_totals = defaultdict(float)

    # Iterar através das linhas para somar amounts
    for row in ws.iter_rows(min_row=2, max_col=8, values_only=True):
        if row[1] == round_votium:  # Coluna B para verificar o round
            gauge_address = (row[7]).lower()  # Coluna H
            amount = row[6]  # Coluna G
            if amount is None or amount == '' or amount == 'ERROR':
                amount = 0
            elif isinstance(amount, str) and amount.startswith('='):
                amount = 0
            else:
                amount = float(amount)

            gauge_totals[gauge_address] += amount

    # Criar ou selecionar a planilha de resultados
    if "Agrupados" not in wb.sheetnames:
        ws_new = wb.create_sheet("Agrupados")
    else:
        ws_new = wb["Agrupados"]

    ws_new.append(["Gauge Address", "Pool Name", "Bribe Total Amount $", "CRV Emissions", 'Value Emissions', 'CRV Price', 'Bribe ROI'])

    # Escrever os dados agrupados com Pool Names
    for gauge, total_bribe in gauge_totals.items():
        pool_name = gauge_to_poolname.get(gauge, "Unknown Pool")
        
        wb_data = get(pool_name, gauge)
        crv_emissions = 0
        crv_price = 0
        for row in wb_data.iter_rows(min_row=2, values_only=True):
            round = row[1]
            
            if round==(round_votium-1)*2 or round==((round_votium-1)*2)+1:
                crv_emissions += int(row[6])

            if round==round_votium*2+1:
                crv_price = row[8]            
            
        dolar_value_emissions = crv_emissions * crv_price

        try:
            bribe_roi = dolar_value_emissions / total_bribe
        except:
            bribe_roi = 'No bribe for this gauge'
        ws_new.append([gauge, pool_name, total_bribe, crv_emissions, dolar_value_emissions, crv_price, bribe_roi])

    # Salvar o workbook modificado
    new_file_path = f"curve/votium_bribe_round/agroupad_spreadsheets/Agroupad_PoolsRound{round_votium}.xlsx"
    wb.save(new_file_path)

